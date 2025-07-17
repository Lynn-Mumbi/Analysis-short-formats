from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
import pandas as pd

input_path = "C:\\Users\\Administrator\\PycharmProjects\\Mix\\MORL night driving\\Detailed Event Report - morl - june.xlsx"
final_output_path = "C:\\Users\\Administrator\\Documents\\night driving menengai occurrence\\Night driving occurrences morl - june.xlsx"

wb = load_workbook("C:\\Users\\Administrator\\PycharmProjects\\Mix\\MORL night driving\\Detailed Event Report - morl - june.xlsx")
ws = wb.active

# remoeing images from sheet
ws._images = []

# unmerge all merged cells
for merged_range in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merged_range))

#there's borders betwwen row 3 and 4 I want out hope it doesnt spoil a thing
'''thin = Side(border_style = None)
empty_border = Border(top = thin, bottom= thin, left = thin, right = thin)

for col in range(1, ws.max_column +1):
    if ws.cell(row=3, column = col).border.bottom:
        ws.cell(row=3, column = col).border = empty_border
    if ws.cell(row = 4, column = col).border.top:
        ws.cell(row = 4, column = col). border = empty_border
'''


# clearing formatting
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font()
        cell.fill = PatternFill()
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'


# setting date and time formats
for row in ws.iter_rows(min_row = 1 , max_row = ws.max_row):
    # column F - start date
    row[5].number_format = 'yyyy-mm-dd'
    # column G - start time
    row[6].number_format = 'hh:mm:ss'
    # column H - end time
    row[7].number_format = 'hh:mm:ss'


# fist 6 rows delete
ws. delete_rows(1,6)






# saving cleaned file
cleaned_path = "C:\\Users\\Administrator\\Documents\\night driving menengai occurrence\\Night driving temp cleaned - june.xlsx"
wb.save(cleaned_path)


#################################33
# loading data into Data Frame
df = pd.read_excel(cleaned_path)

# date is treated as date only
print(df.columns)
df['Start Date'] = pd.to_datetime(df['Start Date']).dt.date
# keeping only first occurrence each day-dropping dups
df = df.drop_duplicates(subset=['Start Date', 'Registration Number'], keep='first')
# sort
df = df.sort_values(by=['Start Date', 'Registration Number']).reset_index( drop = True)

# grouping by reg number and counting how many days they appear
monthly_counts = df.groupby('Registration Number')['Start Date'].nunique().reset_index()

# rename for clarity
monthly_counts.rename(columns = {'Start Date': 'Monthly Occurrence Count'}, inplace = True)


# saving file
df.to_excel(final_output_path, index = False)
# monthly count separate
monthly_counts.to_excel("C:\\Users\\Administrator\\Documents\\night driving menengai occurrence\\Monthly Occurrence Summary - morl - june.xlsx", index=False)

